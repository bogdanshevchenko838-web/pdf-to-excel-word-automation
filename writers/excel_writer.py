from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import OUTPUT_DIR


def save_to_excel(data: dict):
    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    invoice_id = data.get("invoice_number") or "output"
    excel_path = output_dir / f"invoice_{invoice_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice Number", "Order Number", "Invoice Date", "Due Date",
        "From", "To",
        "Hrs/Qty", "Service", "Rate/Price", "Adjust", "Sub Total", "Tax", "Total",
    ]
    ws.append(headers)

    row = [
        data.get("invoice_number", ""),
        data.get("order_number", ""),
        data.get("invoice_date", ""),
        data.get("due_date", ""),
        data.get("from", ""),
        data.get("to", ""),
        data.get("hrs_qty", ""),
        data.get("service", ""),
        data.get("rate_price", ""),
        data.get("adjust", ""),
        data.get("sub_total", ""),
        data.get("tax", ""),
        data.get("total", ""),
    ]
    ws.append(row)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for idx, cell in enumerate(ws[2], start=1):
        cell.border = border
        # From и To — перенос строк
        if idx in (5, 6):  # From, To
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(excel_path)
    print(f"[+] Excel created: {excel_path}")
